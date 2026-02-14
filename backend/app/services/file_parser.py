"""File parsing utilities."""
import csv
import io

try:
    import openpyxl
except ImportError:
    openpyxl = None


def parse_excel_or_csv(content: bytes, filename: str) -> list[list[str]]:
    ext = filename.split(".")[-1].lower() if "." in filename else ""
    if ext == "csv":
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        return [row for row in reader]
    if ext in ("xlsx", "xls") and openpyxl:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        return [[str(c.value or "") for c in row] for row in ws.iter_rows()]
    return []
