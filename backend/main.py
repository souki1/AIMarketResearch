"""
CustomMarket backend - PostgreSQL (users, tabs) + MongoDB (documents).
"""
import csv
import io
import os
import re
import uuid
from bson import Binary
import httpx
from dotenv import load_dotenv
from pymongo import MongoClient, ReturnDocument
from fastapi import FastAPI, File, Form, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from jose import JWTError, jwt
from passlib.context import CryptContext
from pydantic import BaseModel
from datetime import datetime
from sqlalchemy import Column, Integer, String, Text, DateTime, create_engine
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.orm import DeclarativeBase

try:
    import openpyxl
except ImportError:
    openpyxl = None

load_dotenv()

DATABASE_URL = os.getenv(
    "DATABASE_URL", "postgresql+asyncpg://postgres:postgres@localhost:5432/custommarket"
)
SECRET_KEY = os.getenv("SECRET_KEY", "dev-secret-change-in-production")
MONGODB_URI = os.getenv("MONGODB_URI", "mongodb://localhost:27017")
MONGODB_DB = os.getenv("MONGODB_DB", "custommarket")
OLLAMA_URL = os.getenv("OLLAMA_URL", "http://localhost:11434")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "llama2:latest")

# MongoDB - flexible schema for documents
_mongo_client = MongoClient(MONGODB_URI)
_mongo_db = _mongo_client[MONGODB_DB]
documents_coll = _mongo_db["documents"]
counters_coll = _mongo_db["counters"]

# Sync engine for migrations (create tables)
SYNC_URL = DATABASE_URL.replace("+asyncpg", "").replace("postgresql+asyncpg", "postgresql")
engine_sync = create_engine(SYNC_URL)
engine = create_async_engine(DATABASE_URL, echo=False)
async_session = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


class Base(DeclarativeBase):
    pass


class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, autoincrement=True)
    email = Column(String(255), unique=True, nullable=False)
    hashed_password = Column(String(255), nullable=False)
    name = Column(String(255), nullable=True)
    workspace_id = Column(String(36), nullable=False)
    workspace_name = Column(String(255), nullable=False)


class DataTab(Base):
    __tablename__ = "data_tabs"
    id = Column(Integer, primary_key=True, autoincrement=True)
    workspace_id = Column(String(36), nullable=False)
    name = Column(String(255), nullable=False)
    sort_order = Column(Integer, default=0)


class ResearchRequest(Base):
    __tablename__ = "research_requests"
    id = Column(Integer, primary_key=True, autoincrement=True)
    workspace_id = Column(String(36), nullable=False)
    file_id = Column(Integer, nullable=False)
    filename = Column(String(255), nullable=False, default="")
    selected_rows = Column(JSONB, nullable=False, default=list)  # [1, 2, 3]
    selected_columns = Column(JSONB, nullable=False, default=list)  # [0, 1]
    why_fields = Column(Text, nullable=False, default="")
    what_result = Column(Text, nullable=False, default="")
    status = Column(String(50), nullable=False, default="pending")
    created_at = Column(DateTime, nullable=False, default=datetime.utcnow)


class ResearchAllRequest(Base):
    __tablename__ = "research_all_requests"
    id = Column(Integer, primary_key=True, autoincrement=True)
    workspace_id = Column(String(36), nullable=False)
    file_id = Column(Integer, nullable=False)
    filename = Column(String(255), nullable=False, default="")
    total_rows = Column(Integer, nullable=False, default=0)
    total_columns = Column(Integer, nullable=False, default=0)
    why_fields = Column(Text, nullable=False, default="")
    what_result = Column(Text, nullable=False, default="")
    status = Column(String(50), nullable=False, default="pending")
    created_at = Column(DateTime, nullable=False, default=datetime.utcnow)


class AnalyzeQuery(Base):
    __tablename__ = "analyze_query"
    id = Column(Integer, primary_key=True, autoincrement=True)
    workspace_id = Column(String(36), nullable=False)
    research_request_id = Column(Integer, nullable=True)
    research_all_request_id = Column(Integer, nullable=True)
    file_id = Column(Integer, nullable=False)
    filename = Column(String(255), nullable=False, default="")
    query_template = Column(Text, nullable=False, default="")
    selected_columns = Column(JSONB, nullable=False, default=list)
    why_fields = Column(Text, nullable=False, default="")
    what_result = Column(Text, nullable=False, default="")
    row_count = Column(Integer, nullable=False, default=0)
    column_count = Column(Integer, nullable=False, default=0)
    status = Column(String(50), nullable=False, default="pending")
    created_at = Column(DateTime, nullable=False, default=datetime.utcnow)


class SearchableQuery(Base):
    __tablename__ = "searchable_query"
    id = Column(Integer, primary_key=True, autoincrement=True)
    analyze_query_id = Column(Integer, nullable=False)
    workspace_id = Column(String(36), nullable=False)
    row_index = Column(Integer, nullable=False, default=0)
    row_values = Column(JSONB, nullable=False, default=dict)
    query_text = Column(Text, nullable=False, default="")
    status = Column(String(50), nullable=False, default="pending")
    created_at = Column(DateTime, nullable=False, default=datetime.utcnow)


def _init_db():
    try:
        Base.metadata.create_all(engine_sync)
    except Exception as e:
        print(f"PostgreSQL connection failed: {e}")
        print("Create backend/.env with: DATABASE_URL=postgresql+asyncpg://USER:PASSWORD@localhost:5432/aimarket")


def _next_doc_seq() -> int:
    """Get next numeric id for documents (frontend compatibility)."""
    r = counters_coll.find_one_and_update(
        {"_id": "document_seq"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=ReturnDocument.AFTER,
    )
    return r["seq"]


def _to_json_safe(obj):
    """Convert parsed data to MongoDB-safe format (handles any structure)."""
    if obj is None:
        return None
    if isinstance(obj, (list, dict, str, int, float, bool)):
        return obj
    return str(obj)


# Common LLM filler phrases to strip (order matters - longer first)
_LLM_FILLER_PATTERNS = [
    r"^(?:Sure!?|Okay!?|Certainly!?)\s*",
    r"^(?:Based on (?:the )?information provided,?\s*)?(?:here is|here's) (?:a |the )?(?:short )?search query (?:that you can use|below):?\s*",
    r"^(?:You can use (?:the )?following (?:search )?query:?\s*)?",
    r"^(?:The (?:search )?query (?:is|would be):?\s*)?",
    r"^[^\[\w]*(?:query|search):\s*",
    r"^[\[].*$",  # line starting with [ is likely the query
]


def _extract_query_only(raw: str, column_names: list[str]) -> str:
    """Strip LLM filler and return only the searchable query."""
    if not raw:
        return ""
    text = raw.strip().strip('"\'')
    # Remove common filler prefixes from the whole response
    for pat in _LLM_FILLER_PATTERNS:
        text = re.sub(pat, "", text, flags=re.IGNORECASE).strip()
    # Split into lines and find the best candidate
    lines = [ln.strip().strip('"\'') for ln in text.split("\n") if ln.strip()]
    for line in lines:
        # Prefer line that has [placeholder] - that's the query
        if "[" in line and "]" in line:
            return line
        # Or text after colon (e.g. "Query: alternative suppliers [MPN]")
        if ":" in line:
            after = line.split(":", 1)[-1].strip().strip('"\'')
            if after and len(after) < 100:
                return after
    # Fallback: first non-empty line that's short enough to be a query
    for line in lines:
        if line and len(line) < 150:
            return line
    return ""


def _call_ollama_for_query_template(
    column_names: list[str], why_fields: str, what_result: str
) -> str:
    """Call local Ollama LLM to generate a search query. AI must analyze user requirement + field names, then build the query."""
    columns_str = ", ".join(column_names)
    placeholders_example = " ".join(f"[{c}]" for c in column_names)
    prompt = f"""ANALYZE the user's requirement and the selected fields. Then output ONE Google search query TEMPLATE.

IMPORTANT: You output a TEMPLATE with [ColumnName] placeholders. The system will replace each [ColumnName] with the actual ROW VALUE (e.g. [Part Name] becomes "NTN Tapered Roller Bearing", [Manufacturer Part] becomes "4T-30205"). The final search query will have VALUES in quotes, NOT column headers. Never put "Part Name" or headers as literal text—only [ColumnName] format.

STEP 1 - ANALYZE:
- User's requirement: {why_fields or 'Not specified'}
- User's desired result: {what_result or 'Not specified'}
- Selected fields (use as [ColumnName] placeholders): {columns_str}

STEP 2 - BUILD THE TEMPLATE:
Include: (1) search terms from user intent in "quotes", (2) each column as [ColumnName] so it gets replaced with actual values.
Example: "alternative suppliers" "price" "vendor" {placeholders_example}
When filled, this becomes: "alternative suppliers" "price" "vendor" "NTN Tapered Roller Bearing" "Bearings" "NTN" "4T-30205"

Output ONLY the template, nothing else:"""
    try:
        with httpx.Client(timeout=60.0) as client:
            r = client.post(
                f"{OLLAMA_URL.rstrip('/')}/api/generate",
                json={
                    "model": OLLAMA_MODEL,
                    "prompt": prompt,
                    "stream": False,
                    "system": "You output a search query TEMPLATE with [ColumnName] placeholders. Each placeholder will be replaced with the actual row value (e.g. [Part Name] -> 'Siemens Circuit Breaker'). The final query has values in quotes, not headers. Output only the template.",
                },
            )
            r.raise_for_status()
            data = r.json()
            raw = (data.get("response") or "").strip()
            extracted = _extract_query_only(raw, column_names)
            if not extracted:
                raise HTTPException(
                    status_code=503,
                    detail="LLM did not return a valid query. Please try again or rephrase your request.",
                )
            return extracted
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=503,
            detail=f"LLM unavailable: {str(e)}. Ensure Ollama is running and the model is loaded.",
        )


def _fill_query_template(
    template: str, column_names: list[str], row_values: list[str]
) -> str:
    """Replace [ColumnName] placeholders with actual row values. Values are wrapped in " " for precise Google search."""
    result = template
    for col_name, val in zip(column_names, row_values):
        placeholder = f"[{col_name}]"
        raw_val = str(val or "").strip()
        # Wrap in double quotes for precise/exact match in Google search
        quoted_val = f'"{raw_val}"' if raw_val else ""
        result = result.replace(placeholder, quoted_val)
    # Replace any remaining [X] with empty string (in case LLM used different format)
    result = re.sub(r"\[\w[^\]]*\]", "", result)
    return " ".join(result.split())  # normalize whitespace


def _get_parsed_data(doc: dict) -> tuple[list[str], list[list[str]]]:
    """Extract headers and data rows from MongoDB document parsed_data."""
    parsed = doc.get("parsed_data")
    if not parsed or not isinstance(parsed, list):
        return [], []
    rows = [[str(c or "") for c in r] for r in parsed]
    if not rows:
        return [], []
    headers = rows[0]
    data_rows = rows[1:]
    return headers, data_rows


def parse_excel_or_csv(content: bytes, filename: str) -> list[list[str]]:
    ext = filename.split(".")[-1].lower() if "." in filename else ""
    if ext == "csv":
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        return [row for row in reader]
    if ext in ("xlsx", "xls") and openpyxl:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        return [[str(c.value or "") for c in row] for row in ws.iter_rows()]
    return []


def create_access_token(user_id: int) -> str:
    return jwt.encode({"sub": str(user_id)}, SECRET_KEY, algorithm="HS256")


def get_user_id_from_token(token: str) -> int | None:
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        return int(payload.get("sub", 0))
    except (JWTError, ValueError):
        return None


async def get_current_user(authorization: str | None = Header(None, alias="Authorization")) -> User:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = authorization.split(" ", 1)[1]
    user_id = get_user_id_from_token(token)
    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid token")
    async with async_session() as db:
        from sqlalchemy import select
        r = await db.execute(select(User).where(User.id == user_id))
        user = r.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user


app = FastAPI(title="CustomMarket API")


@app.on_event("startup")
def startup():
    _init_db()


app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# --- Auth ---
class LoginRequest(BaseModel):
    email: str
    password: str


class RegisterRequest(BaseModel):
    email: str
    password: str
    name: str | None = None


@app.post("/api/auth/login")
async def login(req: LoginRequest):
    from sqlalchemy import select
    async with async_session() as db:
        r = await db.execute(select(User).where(User.email == req.email))
        user = r.scalar_one_or_none()
    if not user or not pwd_context.verify(req.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    return {"access_token": create_access_token(user.id)}


@app.post("/api/auth/register")
async def register(req: RegisterRequest):
    from sqlalchemy import select
    async with async_session() as db:
        r = await db.execute(select(User).where(User.email == req.email))
        if r.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Email already registered")
        ws_id = str(uuid.uuid4())
        ws_name = (req.name or req.email.split("@")[0]) + "'s Workspace"
        user = User(
            email=req.email,
            hashed_password=pwd_context.hash(req.password),
            name=req.name,
            workspace_id=ws_id,
            workspace_name=ws_name,
        )
        db.add(user)
        await db.commit()
        await db.refresh(user)
    return {
        "id": user.id,
        "email": user.email,
        "name": user.name,
        "workspace_id": user.workspace_id,
        "workspace_name": user.workspace_name,
    }


@app.get("/api/auth/me")
async def auth_me(authorization: str | None = Header(None, alias="Authorization")):
    user = await get_current_user(authorization)
    return {
        "id": user.id,
        "email": user.email,
        "name": user.name,
        "workspace_id": user.workspace_id,
        "workspace_name": user.workspace_name,
    }


# --- Tabs ---
@app.get("/api/tabs")
async def list_tabs(authorization: str | None = Header(None, alias="Authorization")):
    user = await get_current_user(authorization)
    from sqlalchemy import select
    async with async_session() as db:
        r = await db.execute(
            select(DataTab)
            .where(DataTab.workspace_id == user.workspace_id)
            .order_by(DataTab.sort_order, DataTab.id)
        )
        tabs = r.scalars().all()
    tab_ids = [t.id for t in tabs]
    # MongoDB aggregation for file counts per tab
    pipeline = [
        {"$match": {"workspace_id": user.workspace_id, "tab_id": {"$in": tab_ids}}},
        {"$group": {"_id": "$tab_id", "count": {"$sum": 1}}},
    ]
    counts = {r["_id"]: r["count"] for r in documents_coll.aggregate(pipeline)}
    return [
        {"id": t.id, "name": t.name, "sort_order": t.sort_order, "file_count": counts.get(t.id, 0)}
        for t in tabs
    ]


class TabCreate(BaseModel):
    name: str = "New Tab"


class TabRename(BaseModel):
    name: str


@app.post("/api/tabs")
async def create_tab(
    body: TabCreate,
    authorization: str | None = Header(None, alias="Authorization"),
):
    user = await get_current_user(authorization)
    name = body.name
    from sqlalchemy import select, func
    async with async_session() as db:
        r = await db.execute(
            select(func.coalesce(func.max(DataTab.sort_order), 0))
            .where(DataTab.workspace_id == user.workspace_id)
        )
        max_order = (r.scalar() or 0) + 1
        tab = DataTab(workspace_id=user.workspace_id, name=name, sort_order=max_order)
        db.add(tab)
        await db.commit()
        await db.refresh(tab)
    return {"id": tab.id, "name": tab.name, "sort_order": tab.sort_order}


@app.patch("/api/tabs/{tab_id}")
async def rename_tab(tab_id: int, body: TabRename, authorization: str | None = Header(None, alias="Authorization")):
    user = await get_current_user(authorization)
    name = body.name
    from sqlalchemy import select
    async with async_session() as db:
        r = await db.execute(
            select(DataTab).where(
                DataTab.id == tab_id,
                DataTab.workspace_id == user.workspace_id,
            )
        )
        tab = r.scalar_one_or_none()
        if not tab:
            raise HTTPException(status_code=404, detail="Tab not found")
        tab.name = name
        await db.commit()
        await db.refresh(tab)
    return {"id": tab.id, "name": tab.name, "sort_order": tab.sort_order}


# --- Files (MongoDB - flexible schema) ---
@app.get("/api/files")
async def list_files(tab_id: int | None = None, authorization: str | None = Header(None, alias="Authorization")):
    user = await get_current_user(authorization)
    query = {"workspace_id": user.workspace_id}
    if tab_id is not None:
        query["tab_id"] = tab_id
    cursor = documents_coll.find(query).sort("id", 1)
    files = []
    for doc in cursor:
        files.append({
            "id": doc["id"],
            "document_id": doc.get("document_id", ""),
            "filename": doc["filename"],
            "storage_path": "",
            "mime_type": doc.get("mime_type"),
            "size": doc.get("size"),
            "tab_id": doc.get("tab_id"),
            "parsed_data": doc.get("parsed_data"),
            "notes": doc.get("notes", ""),
        })
    return files


@app.post("/api/files/upload")
async def upload_files(
    files: list[UploadFile] = File(alias="files"),
    tab_id: int | None = Form(None),
    authorization: str | None = Header(None, alias="Authorization"),
):
    user = await get_current_user(authorization)
    uploaded = []
    for uf in files:
        content = await uf.read()
        ext = (uf.filename or "").split(".")[-1].lower()
        parsed = parse_excel_or_csv(content, uf.filename or "") if ext in ("csv", "xlsx", "xls") else None
        doc_id = str(uuid.uuid4())
        seq_id = _next_doc_seq()
        parsed_safe = _to_json_safe(parsed) if parsed is not None else None
        doc = {
            "id": seq_id,
            "document_id": doc_id,
            "workspace_id": user.workspace_id,
            "tab_id": tab_id,
            "filename": uf.filename or "file",
            "mime_type": uf.content_type,
            "size": len(content),
            "file_content": Binary(content),
            "parsed_data": parsed_safe,
        }
        documents_coll.insert_one(doc)
        uploaded.append({
            "id": seq_id,
            "document_id": doc_id,
            "filename": doc["filename"],
            "storage_path": "",
            "mime_type": doc["mime_type"],
            "size": doc["size"],
            "tab_id": tab_id,
            "parsed_data": parsed_safe,
        })
    return {"uploaded": uploaded}


@app.get("/api/files/{file_id}/content")
async def get_file_content(
    file_id: int,
    authorization: str | None = Header(None, alias="Authorization"),
    token: str | None = None,
):
    auth = authorization or (f"Bearer {token}" if token else None)
    user = await get_current_user(auth)
    doc = documents_coll.find_one({"id": file_id, "workspace_id": user.workspace_id})
    if not doc:
        raise HTTPException(status_code=404, detail="File not found")
    content = doc.get("file_content")
    if not content:
        raise HTTPException(status_code=404, detail="File content not found")
    from fastapi.responses import Response
    return Response(
        content=bytes(content),
        media_type=doc.get("mime_type") or "application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{doc["filename"]}"'},
    )


class FileNotesUpdate(BaseModel):
    notes: str = ""


class ResearchRequestCreate(BaseModel):
    file_id: int
    selected_rows: list[int] = []
    selected_columns: list[int] = []
    why_fields: str = ""
    what_result: str = ""


class ResearchAllRequestCreate(BaseModel):
    file_id: int
    total_rows: int = 0
    total_columns: int = 0
    why_fields: str = ""
    what_result: str = ""


@app.patch("/api/files/{file_id}")
async def update_file_notes(
    file_id: int,
    body: FileNotesUpdate,
    authorization: str | None = Header(None, alias="Authorization"),
):
    user = await get_current_user(authorization)
    r = documents_coll.update_one(
        {"id": file_id, "workspace_id": user.workspace_id},
        {"$set": {"notes": body.notes}},
    )
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="File not found")
    return {"ok": True, "notes": body.notes}


@app.delete("/api/files/{file_id}")
async def delete_file(file_id: int, authorization: str | None = Header(None, alias="Authorization")):
    user = await get_current_user(authorization)
    r = documents_coll.delete_one({"id": file_id, "workspace_id": user.workspace_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="File not found")
    return {"ok": True}


# --- Research Requests (PostgreSQL) ---
@app.post("/api/research-requests")
async def create_research_request(
    body: ResearchRequestCreate,
    authorization: str | None = Header(None, alias="Authorization"),
):
    user = await get_current_user(authorization)
    doc = documents_coll.find_one({"id": body.file_id, "workspace_id": user.workspace_id})
    if not doc:
        raise HTTPException(status_code=404, detail="File not found")
    filename = doc.get("filename", "")
    headers, data_rows = _get_parsed_data(doc)
    if not headers or not data_rows:
        raise HTTPException(status_code=400, detail="File has no tabular data")

    col_indices = [i for i in body.selected_columns if 0 <= i < len(headers)]
    row_indices = [i for i in body.selected_rows if 0 <= i < len(data_rows)]
    if not col_indices:
        col_indices = list(range(len(headers)))
    if not row_indices:
        row_indices = list(range(len(data_rows)))

    column_names = [headers[i] for i in col_indices]
    rows_to_process = [data_rows[i] for i in row_indices]

    query_template = _call_ollama_for_query_template(
        column_names, body.why_fields, body.what_result
    )

    async with async_session() as db:
        req = ResearchRequest(
            workspace_id=user.workspace_id,
            file_id=body.file_id,
            filename=filename,
            selected_rows=body.selected_rows,
            selected_columns=body.selected_columns,
            why_fields=body.why_fields,
            what_result=body.what_result,
            status="pending",
        )
        db.add(req)
        await db.commit()
        await db.refresh(req)

        analyze = AnalyzeQuery(
            workspace_id=user.workspace_id,
            research_request_id=req.id,
            research_all_request_id=None,
            file_id=body.file_id,
            filename=filename,
            query_template=query_template,
            selected_columns=column_names,
            why_fields=body.why_fields,
            what_result=body.what_result,
            row_count=len(rows_to_process),
            column_count=len(column_names),
            status="completed",
        )
        db.add(analyze)
        await db.commit()
        await db.refresh(analyze)

        for row_idx, row in enumerate(rows_to_process):
            row_vals = [row[c] if c < len(row) else "" for c in col_indices]
            row_values_dict = dict(zip(column_names, row_vals))
            query_text = _fill_query_template(query_template, column_names, row_vals)
            sq = SearchableQuery(
                analyze_query_id=analyze.id,
                workspace_id=user.workspace_id,
                row_index=row_idx,
                row_values=row_values_dict,
                query_text=query_text,
                status="pending",
            )
            db.add(sq)
        await db.commit()

    return {"id": req.id, "analyze_id": analyze.id, "searchable_count": len(rows_to_process), "ok": True}


# --- Research All Requests (PostgreSQL - separate table) ---
@app.post("/api/research-all-requests")
async def create_research_all_request(
    body: ResearchAllRequestCreate,
    authorization: str | None = Header(None, alias="Authorization"),
):
    user = await get_current_user(authorization)
    doc = documents_coll.find_one({"id": body.file_id, "workspace_id": user.workspace_id})
    if not doc:
        raise HTTPException(status_code=404, detail="File not found")
    filename = doc.get("filename", "")
    headers, data_rows = _get_parsed_data(doc)
    if not headers or not data_rows:
        raise HTTPException(status_code=400, detail="File has no tabular data")

    col_count = min(body.total_columns, len(headers)) or len(headers)
    row_count = min(body.total_rows, len(data_rows)) or len(data_rows)
    col_indices = list(range(col_count))
    row_indices = list(range(row_count))

    column_names = [headers[i] for i in col_indices]
    rows_to_process = [data_rows[i] for i in row_indices]

    query_template = _call_ollama_for_query_template(
        column_names, body.why_fields, body.what_result
    )

    async with async_session() as db:
        req = ResearchAllRequest(
            workspace_id=user.workspace_id,
            file_id=body.file_id,
            filename=filename,
            total_rows=row_count,
            total_columns=col_count,
            why_fields=body.why_fields,
            what_result=body.what_result,
            status="pending",
        )
        db.add(req)
        await db.commit()
        await db.refresh(req)

        analyze = AnalyzeQuery(
            workspace_id=user.workspace_id,
            research_request_id=None,
            research_all_request_id=req.id,
            file_id=body.file_id,
            filename=filename,
            query_template=query_template,
            selected_columns=column_names,
            why_fields=body.why_fields,
            what_result=body.what_result,
            row_count=len(rows_to_process),
            column_count=len(column_names),
            status="completed",
        )
        db.add(analyze)
        await db.commit()
        await db.refresh(analyze)

        for row_idx, row in enumerate(rows_to_process):
            row_vals = [row[c] if c < len(row) else "" for c in col_indices]
            row_values_dict = dict(zip(column_names, row_vals))
            query_text = _fill_query_template(query_template, column_names, row_vals)
            sq = SearchableQuery(
                analyze_query_id=analyze.id,
                workspace_id=user.workspace_id,
                row_index=row_idx,
                row_values=row_values_dict,
                query_text=query_text,
                status="pending",
            )
            db.add(sq)
        await db.commit()

    return {"id": req.id, "analyze_id": analyze.id, "searchable_count": len(rows_to_process), "ok": True}


@app.get("/health")
async def health():
    return {"status": "ok"}


@app.get("/health/db")
async def health_db():
    from sqlalchemy import text
    result = {}
    try:
        async with engine.connect() as conn:
            await conn.execute(text("SELECT 1"))
        result["postgresql"] = "connected"
    except Exception as e:
        result["postgresql"] = str(e)
    try:
        _mongo_client.admin.command("ping")
        result["mongodb"] = "connected"
    except Exception as e:
        result["mongodb"] = str(e)
    if result.get("postgresql") != "connected" or result.get("mongodb") != "connected":
        raise HTTPException(status_code=503, detail=result)
    return {"status": "ok", **result}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
